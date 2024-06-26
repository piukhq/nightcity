import csv
import json
from io import BytesIO, StringIO

import openpyxl
import pendulum
import requests
from box import Box
from pydantic import EmailStr
from pydantic_settings import BaseSettings, SettingsConfigDict
from requests.auth import HTTPBasicAuth
from sqlalchemy.engine.row import Row
from tenacity import retry, stop_after_attempt, wait_fixed

from nightcity.azure import keyvault_client, sftp_client
from nightcity.screamsheet.query import get_marketing_data, get_transaction_data
from nightcity.settings import log


class ScreamsheetConfig(BaseSettings):
    """Config for Screamsheet."""

    mailgun_from: EmailStr = "noreply@bink.com"
    mailgun_to: EmailStr = "onlineservices@bink.com"

    model_config = SettingsConfigDict(env_prefix="SCREAMSHEET_")


app_settings = ScreamsheetConfig()


def prepare_csv(data: list[Row], headers: list[str]) -> StringIO:
    """Write Marketing Preferences to a CSV StringIO Object."""
    f = StringIO()
    file_writer = csv.writer(f)
    file_writer.writerow(headers)
    file_writer.writerows(data)
    return f


def prepare_xlsx(data: list[Row], headers: list[str]) -> StringIO:
    """Write Marketing Preferences to a XLSX StringIO Object."""
    f = BytesIO()
    wb = openpyxl.Workbook()
    sheet = wb.active
    for colno, heading in enumerate(headers, 1):
        sheet.cell(row=1, value=heading, column=colno)
    for rowno, row in enumerate(data, 2):
        for colno, value in enumerate(row, 1):
            sheet.cell(row=rowno, column=colno, value=value)
    wb.save(f)
    return f


def generate_filename(data_type: str, folder: str, ext: str | None = None) -> str:
    """Generate a Filename for Azure Storage."""
    date = pendulum.now(tz="utc").strftime("%Y-%m-%d_%H-%M-%S")
    return f"{folder}/{data_type}_{date}.{ext}" if ext else f"{folder}/{data_type}_{date}.csv"


def upload_blob(file: StringIO, filename: str) -> None:
    """Upload Data to Azure Storage."""
    blob = sftp_client.get_blob_client(
        container="viator",
        blob=filename,
    )
    log.info(f"Uploading {filename} to Azure Storage")
    blob.upload_blob(file.read())


@retry(stop=stop_after_attempt(3), wait=wait_fixed(10))
def send_email_via_mailgun(message: str, subject: str) -> None:
    """Send an email via Mailgun."""
    mailgun_secret: Box = Box(json.loads(keyvault_client.get_secret("mailgun").value))

    if mailgun_secret.MAILGUN_API_KEY == "fake-api-key":
        log.error("Mailgun API Key is not set")
        return

    log.info("Sending email via Mailgun")
    try:
        r = requests.post(
            f"{mailgun_secret.MAILGUN_API}/{mailgun_secret.MAILGUN_DOMAIN}/messages",
            auth=HTTPBasicAuth("api", mailgun_secret.MAILGUN_API_KEY),
            data={
                "from": str(app_settings.mailgun_from),
                "to": str(app_settings.mailgun_to),
                "subject": subject,
                "text": message,
            },
            timeout=10,
        )
        r.raise_for_status()
        log.info("Mailgun Request Successful")
    except requests.RequestException:
        log.exception("Mailgun Request Failed")


def send_marketing_info() -> None:
    """Entrypoint for Screamsheet."""
    marketing_data = get_marketing_data()
    file = prepare_csv(
        marketing_data,
        headers=[
            "Email",
            "First Name",
            "Last Name",
            "Date of Birth",
            "Key Name",
            "Value",
            "Account Number",
            "Updated At",
        ],
    )
    file.seek(0)
    filename = generate_filename("Marketing_Preferences", "downloads")
    upload_blob(file=file, filename=filename)
    subject = "Viator Marketing Preferences"
    message = """
    Hi Viator Team,

    Please note the monthly Viator Marketing preference file has been uploaded to the SFTP

    Thanks
    Bink Team
    """

    send_email_via_mailgun(message=message, subject=subject)
    url = "https://ping.checklyhq.com/17445d79-4bf3-4956-bbe4-86a3c06c940b"
    requests.get(url, timeout=5)


def send_transcation_info() -> None:
    """Get transaction data from the Harmonia Database."""
    transcation_data = get_transaction_data()
    file = prepare_xlsx(
        transcation_data,
        headers=[
            "AMOUNT",
            "TRANSACTION DATE",
            "AUTH",
            "MID",
            "LAST FOUR",
        ],
    )
    file.seek(0)
    filename = generate_filename("Transactions/viator_weekly_transaction", "downloads", "xlsx")
    upload_blob(file=file, filename=filename)
    subject = "Viator Discounts Weekly Transactions"
    message = """
    Hi Viator team,

    Please see the this weeks Viator transactions file on the SFTP in the Transactions folder.

    Kind Regards
    Bink
    """
    send_email_via_mailgun(message=message, subject=subject)
    url = "https://ping.checklyhq.com/b47dc444-826b-4852-8c19-5a885620f730"
    requests.get(url, timeout=5)
